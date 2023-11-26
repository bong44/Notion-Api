from openpyxl import load_workbook

# 엑셀 파일 경로
excel_file_path = 'min.xlsx'

# 엑셀 파일 로드
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# 1번째 행부터 5번째 행까지 삭제
for _ in range(5):
    sheet.delete_rows(1)

# 배열 초기화
result_array = []

# 엑셀 데이터를 배열에 json 형태로 담기
for row in sheet.iter_rows(min_row=1, values_only=True):
    data = {
        'time': row[0],
        'name': row[2],  # C번째 열
        'amt': row[5]  # F번째 열
    }
    result_array.append(data)

# amt가 1인 데이터만 추출
filtered_array = [item for item in result_array if item['amt'] == 31400]

# 'time'을 기준으로 정렬
sorted_array = sorted(filtered_array, key=lambda x: x['time'])

print("\namt가 1인 데이터:")
print(sorted_array)
